from openpyxl import load_workbook

wb = load_workbook('test_data.xlsx', read_only=True)
data = wb.active

for row in data.iter_rows(max_col=1, max_row=2):
    for cell in row:
        print(cell.value)
