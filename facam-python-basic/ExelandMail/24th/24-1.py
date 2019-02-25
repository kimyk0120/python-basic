'''


'''


from openpyxl import Workbook

wb = Workbook()
# ws = wb.active
ws = wb.create_sheet("sheet_test2")

ws['A1'] = 'test1'
ws['B1'] = 'test2'

wb.save('result.xlsx')







